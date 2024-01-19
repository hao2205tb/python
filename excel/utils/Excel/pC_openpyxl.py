import openpyxl
from openpyxl.utils import get_column_letter

class XlSheet:
    def __init__(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        self.ws = ws 
    
    @property
    def num_rows(self):
        return self.ws.max_row
    
    @property
    def num_cols(self):
        return self.ws.max_column
    
    @property
    def row_indexs(self):
        return range(1, self.num_rows + 1)
    
    @property
    def col_indexs(self):
        return range(1, self.num_cols + 1)

    def get_cell(self, row, col):
        return self.ws.cell(row, col).value
    
    def set_cell(self, row, col, value):
        """using cell to write data
            ex: sheet.cell(row=1, column=1, value="20") -> cell: 20
        """
        self.ws.cell(row, col, value)
    
    def set_cell2(self, cellindex:str, value):
        """using cell index to write data
            ex: sheet[A1]="20" -> cell: '20
        """
        self.ws[cellindex] = value
    
    def set_cell3(self, row, col, value):
        """convert row, col to cellindex and write cell value by set_cell2
        """
        cellindex = XlFile.rowcol2cellindex(row, col)
        self.set_cell2(cellindex, value)
    
    def get_data(self):
        """
            return 2-dim array of sheet
            type: list<list>
        """
        data = list()
        for row in self.ws.iter_rows(values_only=True):
            data.append(list(row))
        return data
    
    
class XlFile:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
    
    def get_sheet(self, sheetname: str = None) -> XlSheet:
        """return sheet
        Args:
            sheetname (str): Default first sheet of workbook
        Returns:
            XlSheet
        """
        if sheetname is None:
            ws = self.wb.active
        else:
            ws = self.wb[sheetname]
        return XlSheet(ws)
    
    @staticmethod
    def rowcol2cellindex(row, col) -> str:
        col = get_column_letter(col)
        return f"{col}{row}"
    
    def save(self, filename:str = None):
        if filename is None:
            filename = self.file_path
        self.wb.save(filename)


if __name__ == "__main__":
    filepath = "data/test.xlsx"
    xl = XlFile(filepath)
    sheet = xl.get_sheet()
    print(sheet.num_cols, sheet.num_rows, list(sheet.col_indexs), sheet.row_indexs)
    print(sheet.get_cell(2, 2))
    sheet.set_cell3(1, 2, '-6')
    xl.save()
    data = sheet.get_data()
    for row in data:
        print(row)

